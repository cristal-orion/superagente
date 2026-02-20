from __future__ import annotations

import json
import os
from io import BytesIO
from pathlib import Path

from fastapi import FastAPI, UploadFile, Form, HTTPException, Depends
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import JSONResponse, StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from fastapi.staticfiles import StaticFiles
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

from calculator import calc_response
from models import CalcRequest, CalcResponse

app = FastAPI(title="PV Sales Calculator")

# Datasheets configuration (use env var for Docker, fallback for local dev)
DATASHEETS_DIR = Path(os.environ.get("DATASHEETS_DIR", Path(__file__).parent / "datasheets"))
VALID_CATEGORIES = ["pannelli", "inverter", "batterie", "pompe", "altro"]
MAX_FILE_SIZE = 10 * 1024 * 1024  # 10MB

# Catalog configuration (use env var for Docker, fallback for local dev)
CATALOG_PATH = Path(os.environ.get("CATALOG_PATH", Path(__file__).parent.parent / "frontend" / "catalog.json"))
# Ensure datasheets directories exist
for category in VALID_CATEGORIES:
    (DATASHEETS_DIR / category).mkdir(parents=True, exist_ok=True)

# Mount static files for serving datasheets
app.mount("/datasheets-files", StaticFiles(directory=str(DATASHEETS_DIR)), name="datasheets-files")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=False,
    allow_methods=["*"],
    allow_headers=["*"],
)


# ─── Startup: init DB and seed ────────────────────────────────────────────────

@app.on_event("startup")
def on_startup() -> None:
    from database import init_db, SessionLocal, engine
    from auth import seed_superadmin
    from sqlalchemy import text
    init_db()
    # Migration: add listino_data column if it doesn't exist yet
    try:
        with engine.connect() as conn:
            conn.execute(text("ALTER TABLE listini ADD COLUMN listino_data TEXT"))
            conn.commit()
    except Exception:
        pass  # Column already exists
    # Migration: add plain_password column to users if it doesn't exist yet
    try:
        with engine.connect() as conn:
            conn.execute(text("ALTER TABLE users ADD COLUMN plain_password TEXT"))
            conn.commit()
    except Exception:
        pass  # Column already exists
    db = SessionLocal()
    try:
        seed_superadmin(db)
    finally:
        db.close()


# ─── Include routers ──────────────────────────────────────────────────────────

from routes.auth import router as auth_router
from routes.users import router as users_router
from routes.listini import router as listini_router

app.include_router(auth_router)
app.include_router(users_router)
app.include_router(listini_router)


# ─── Catalog helpers ──────────────────────────────────────────────────────────

def load_catalog_json() -> dict:
    if CATALOG_PATH.exists():
        with open(CATALOG_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {"version": "default", "items": []}


# ─── Core endpoints ───────────────────────────────────────────────────────────

@app.get("/health")
def health() -> dict[str, str]:
    return {"status": "ok"}


@app.post("/calc", response_model=CalcResponse)
def calc(request: CalcRequest) -> CalcResponse:
    return calc_response(request)


# ─── Catalog user endpoint ────────────────────────────────────────────────────

_optional_bearer = HTTPBearer(auto_error=False)


@app.get("/catalog/me")
async def catalog_for_user(
    credentials: HTTPAuthorizationCredentials | None = Depends(_optional_bearer),
):
    """Return catalog filtered by user's assigned listini, or default if no auth."""
    if not credentials:
        return JSONResponse(load_catalog_json())

    try:
        from database import SessionLocal
        from auth import decode_token
        from models_db import User, ListinoAccess, Listino
        from jose import JWTError

        payload = decode_token(credentials.credentials)
        user_id = int(payload.get("sub", 0))

        db = SessionLocal()
        try:
            user = db.query(User).filter(User.id == user_id, User.is_active == True).first()
            if not user:
                return JSONResponse(load_catalog_json())

            items = []
            seen_ids: set[str] = set()

            # All users (superadmin and agents) see only assigned listini
            assigned = db.query(ListinoAccess).filter(ListinoAccess.user_id == user_id).all()
            for access in assigned:
                listino = db.query(Listino).filter(Listino.id == access.listino_id).first()
                if listino:
                    for item in listino.get_items():
                        item_id = item.get("id", "")
                        if item_id not in seen_ids:
                            seen_ids.add(item_id)
                            items.append(item)

            if not items:
                return JSONResponse(load_catalog_json())

            return JSONResponse({"version": "user-custom", "items": items})
        finally:
            db.close()

    except Exception:
        return JSONResponse(load_catalog_json())


# ─── Populate listino from catalog.json ────────────────────────────────────────

from sqlalchemy.orm import Session as _Session
from database import get_db as _get_db
from auth import require_superadmin as _require_superadmin


@app.post("/api/listini/{listino_id}/populate-from-catalog")
async def populate_listino_from_catalog(
    listino_id: int,
    db: _Session = Depends(_get_db),
    _admin=Depends(_require_superadmin),
) -> dict:
    """Copy all items from catalog.json into catalog_items AND listino_data (editor sections)."""
    import uuid as _uuid_mod
    from collections import OrderedDict as _OD
    from models_db import Listino as _Listino
    listino = db.query(_Listino).filter(_Listino.id == listino_id).first()
    if not listino:
        raise HTTPException(status_code=404, detail="Listino non trovato")
    catalog = load_catalog_json()
    items = catalog.get("items", [])
    if not items:
        raise HTTPException(status_code=400, detail="Il catalogo corrente è vuoto")

    # 1. Save catalog_items (used by calculator for agents)
    listino.set_items(items)

    # 2. Build listino_data sections (used by the visual editor)
    categories: dict = _OD()
    for item in items:
        cat = item.get("category", "Altro")
        if cat not in categories:
            categories[cat] = []
        categories[cat].append(item)

    sections = []
    for cat_title, cat_items in categories.items():
        columns = ["Configurazione", "kW", "kWh acc.", "Prezzo €"]

        rows = []
        for item in cat_items:
            row = [
                item.get("label", ""),
                str(item.get("potenza_kw", "")),
                str(int(item.get("accumulo_kwh", 0)) if item.get("accumulo_kwh", 0) == int(item.get("accumulo_kwh", 0)) else item.get("accumulo_kwh", 0)),
                str(int(item.get("prezzo_eur", 0)) if item.get("prezzo_eur", 0) == int(item.get("prezzo_eur", 0)) else item.get("prezzo_eur", 0)),
            ]
            rows.append(row)

        sections.append({
            "id": "sec" + _uuid_mod.uuid4().hex[:8],
            "title": cat_title,
            "notes": "",
            "columns": columns,
            "rows": rows,
        })

    listino.set_listino_data({"sections": sections})
    db.commit()
    return {"status": "ok", "items_copied": len(items)}


# ═══════════════════════════════════════════════════════════════════════════
# Datasheets Management Endpoints
# ═══════════════════════════════════════════════════════════════════════════

@app.get("/datasheets")
def list_datasheets() -> dict[str, list[dict]]:
    """List all datasheets organized by category."""
    result = {}
    for category in VALID_CATEGORIES:
        category_path = DATASHEETS_DIR / category
        files = []
        if category_path.exists():
            for file in category_path.iterdir():
                if file.is_file() and file.suffix.lower() == ".pdf":
                    files.append({
                        "name": file.stem,
                        "filename": file.name,
                        "category": category,
                        "size": file.stat().st_size,
                        "url": f"/datasheets-files/{category}/{file.name}"
                    })
        # Sort by name
        files.sort(key=lambda x: x["name"].lower())
        result[category] = files
    return result


@app.post("/datasheets/upload")
async def upload_datasheet(
    file: UploadFile,
    category: str = Form(...)
) -> dict:
    """Upload a PDF datasheet to a specific category."""
    # Validate category
    if category not in VALID_CATEGORIES:
        raise HTTPException(
            status_code=400,
            detail=f"Invalid category. Must be one of: {', '.join(VALID_CATEGORIES)}"
        )

    # Validate file type
    if not file.filename or not file.filename.lower().endswith(".pdf"):
        raise HTTPException(
            status_code=400,
            detail="Only PDF files are allowed"
        )

    # Read file content
    content = await file.read()

    # Validate file size
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(
            status_code=400,
            detail=f"File too large. Maximum size is {MAX_FILE_SIZE // (1024*1024)}MB"
        )

    # Sanitize filename
    safe_filename = "".join(c for c in file.filename if c.isalnum() or c in "._- ").strip()
    if not safe_filename.lower().endswith(".pdf"):
        safe_filename += ".pdf"

    # Handle duplicate filenames
    target_path = DATASHEETS_DIR / category / safe_filename
    if target_path.exists():
        # Add timestamp to make unique
        import time
        name_part = safe_filename[:-4]  # Remove .pdf
        safe_filename = f"{name_part}_{int(time.time())}.pdf"
        target_path = DATASHEETS_DIR / category / safe_filename

    # Save file
    with open(target_path, "wb") as f:
        f.write(content)

    return {
        "status": "success",
        "filename": safe_filename,
        "category": category,
        "size": len(content),
        "url": f"/datasheets-files/{category}/{safe_filename}"
    }


@app.delete("/datasheets/{category}/{filename}")
def delete_datasheet(category: str, filename: str) -> dict:
    """Delete a datasheet."""
    # Validate category
    if category not in VALID_CATEGORIES:
        raise HTTPException(status_code=400, detail="Invalid category")

    # Validate filename (prevent path traversal)
    if ".." in filename or "/" in filename or "\\" in filename:
        raise HTTPException(status_code=400, detail="Invalid filename")

    file_path = DATASHEETS_DIR / category / filename

    if not file_path.exists():
        raise HTTPException(status_code=404, detail="File not found")

    if not file_path.is_file():
        raise HTTPException(status_code=400, detail="Not a file")

    os.remove(file_path)

    return {"status": "deleted", "filename": filename, "category": category}


# ═══════════════════════════════════════════════════════════════════════════
# Catalog Excel Export/Import Endpoints
# ═══════════════════════════════════════════════════════════════════════════

def _get_user_catalog_items(credentials) -> tuple[list[dict], str]:
    """Return (items, listino_name) for the authenticated user's assigned listini.

    Falls back to catalog.json if no auth or no assigned listini.
    """
    if not credentials:
        catalog = load_catalog_json()
        return catalog.get("items", []), catalog.get("version", "export")

    try:
        from database import SessionLocal
        from auth import decode_token
        from models_db import User, ListinoAccess, Listino

        payload = decode_token(credentials.credentials)
        user_id = int(payload.get("sub", 0))

        db = SessionLocal()
        try:
            user = db.query(User).filter(User.id == user_id, User.is_active == True).first()
            if not user:
                catalog = load_catalog_json()
                return catalog.get("items", []), catalog.get("version", "export")

            items: list[dict] = []
            seen_ids: set[str] = set()
            listino_name = ""

            assigned = db.query(ListinoAccess).filter(ListinoAccess.user_id == user_id).all()
            for access in assigned:
                listino = db.query(Listino).filter(Listino.id == access.listino_id).first()
                if listino:
                    if not listino_name:
                        listino_name = listino.name
                    for item in listino.get_items():
                        item_id = item.get("id", "")
                        if item_id not in seen_ids:
                            seen_ids.add(item_id)
                            items.append(item)

            if not items:
                catalog = load_catalog_json()
                return catalog.get("items", []), catalog.get("version", "export")

            return items, listino_name or "listino"
        finally:
            db.close()
    except Exception:
        catalog = load_catalog_json()
        return catalog.get("items", []), catalog.get("version", "export")


@app.get("/catalog/export")
def export_catalog_excel(
    credentials: HTTPAuthorizationCredentials | None = Depends(_optional_bearer),
):
    """Export the user's assigned listino as an Excel file."""
    items, listino_name = _get_user_catalog_items(credentials)

    if not items:
        raise HTTPException(status_code=404, detail="Nessun listino trovato")

    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Listino"

    # Define headers
    headers = [
        "ID", "Categoria", "Label", "Potenza (kW)", "Accumulo (kWh)", "Prezzo (EUR)"
    ]

    # Style settings
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="C41E3A", end_color="C41E3A", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Write headers with styling
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Write data rows
    for row_idx, item in enumerate(items, 2):
        row_data = [
            item.get("id", ""),
            item.get("category", ""),
            item.get("label", ""),
            item.get("potenza_kw", 0),
            item.get("accumulo_kwh", 0),
            item.get("prezzo_eur", 0),
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = thin_border
            if col >= 4:  # Numeric columns
                cell.alignment = Alignment(horizontal="right")

    # Set column widths
    column_widths = [20, 40, 20, 12, 14, 14]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col).column_letter].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to bytes
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    safe_name = listino_name.replace(" ", "_")
    filename = f"listino_{safe_name}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.get("/catalog/export-pdf")
def export_catalog_pdf(
    credentials: HTTPAuthorizationCredentials | None = Depends(_optional_bearer),
):
    """Export the user's assigned listino as a PDF file."""
    from fpdf import FPDF

    items, listino_name = _get_user_catalog_items(credentials)

    if not items:
        raise HTTPException(status_code=404, detail="Nessun listino trovato")

    pdf = FPDF(orientation="L", unit="mm", format="A4")
    pdf.set_auto_page_break(auto=True, margin=15)
    pdf.add_page()

    # Title
    pdf.set_font("Helvetica", "B", 16)
    pdf.set_text_color(196, 30, 58)  # brand red
    pdf.cell(0, 12, f"Listino: {listino_name}", new_x="LMARGIN", new_y="NEXT")
    pdf.ln(4)

    # Table header
    col_widths = [50, 70, 55, 30, 35, 35]
    col_headers = ["ID", "Categoria", "Label", "kW", "Accumulo kWh", "Prezzo EUR"]

    pdf.set_font("Helvetica", "B", 10)
    pdf.set_fill_color(196, 30, 58)
    pdf.set_text_color(255, 255, 255)
    for i, header in enumerate(col_headers):
        pdf.cell(col_widths[i], 9, header, border=1, fill=True, align="C")
    pdf.ln()

    # Table rows
    pdf.set_font("Helvetica", "", 9)
    pdf.set_text_color(30, 30, 30)
    fill = False
    for item in items:
        if pdf.get_y() > 180:
            pdf.add_page()
            # Re-draw header on new page
            pdf.set_font("Helvetica", "B", 10)
            pdf.set_fill_color(196, 30, 58)
            pdf.set_text_color(255, 255, 255)
            for i, header in enumerate(col_headers):
                pdf.cell(col_widths[i], 9, header, border=1, fill=True, align="C")
            pdf.ln()
            pdf.set_font("Helvetica", "", 9)
            pdf.set_text_color(30, 30, 30)
            fill = False

        if fill:
            pdf.set_fill_color(245, 245, 245)
        else:
            pdf.set_fill_color(255, 255, 255)

        potenza = item.get("potenza_kw", 0)
        accumulo = item.get("accumulo_kwh", 0)
        prezzo = item.get("prezzo_eur", 0)

        pdf.cell(col_widths[0], 8, str(item.get("id", "")), border=1, fill=True)
        pdf.cell(col_widths[1], 8, str(item.get("category", "")), border=1, fill=True)
        pdf.cell(col_widths[2], 8, str(item.get("label", "")), border=1, fill=True)
        pdf.cell(col_widths[3], 8, f"{potenza:.1f}", border=1, fill=True, align="R")
        pdf.cell(col_widths[4], 8, f"{accumulo:.1f}", border=1, fill=True, align="R")
        pdf.cell(col_widths[5], 8, f"{prezzo:,.0f}", border=1, fill=True, align="R")
        pdf.ln()
        fill = not fill

    # Footer
    pdf.ln(6)
    pdf.set_font("Helvetica", "I", 8)
    pdf.set_text_color(120, 120, 120)
    pdf.cell(0, 6, f"Totale prodotti: {len(items)}", new_x="LMARGIN", new_y="NEXT")

    output = BytesIO()
    pdf.output(output)
    output.seek(0)

    safe_name = listino_name.replace(" ", "_")
    filename = f"listino_{safe_name}.pdf"

    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@app.post("/catalog/import")
async def import_catalog_excel(file: UploadFile):
    """Import an Excel file to update the catalog."""
    # Validate file type
    if not file.filename or not file.filename.lower().endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Only .xlsx files are allowed")

    # Read file
    content = await file.read()

    try:
        wb = load_workbook(BytesIO(content))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Invalid Excel file: {str(e)}")

    # Find the data sheet
    if "Listino" not in wb.sheetnames:
        raise HTTPException(status_code=400, detail="Sheet 'Listino' not found in Excel file")

    ws = wb["Listino"]

    # Read existing catalog for metadata
    if CATALOG_PATH.exists():
        with open(CATALOG_PATH, "r", encoding="utf-8") as f:
            existing_catalog = json.load(f)
    else:
        existing_catalog = {"version": "imported", "source_pdf": ""}

    # Parse header row to get column indices
    headers = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=1, column=col).value
        if header:
            headers[header.strip().lower()] = col

    # Map expected headers to column indices
    col_map = {
        "id": headers.get("id"),
        "category": headers.get("categoria"),
        "label": headers.get("label"),
        "potenza_kw": headers.get("potenza (kw)"),
        "accumulo_kwh": headers.get("accumulo (kwh)"),
        "prezzo_eur": headers.get("prezzo (eur)"),
    }

    # Validate required columns exist
    required = ["id", "category", "label", "potenza_kw", "prezzo_eur"]
    missing = [col for col in required if not col_map.get(col)]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Missing required columns: {', '.join(missing)}"
        )

    # Parse items
    items = []
    errors = []

    for row in range(2, ws.max_row + 1):
        # Skip empty rows
        id_val = ws.cell(row=row, column=col_map["id"]).value
        if not id_val:
            continue

        try:
            # Build item
            item = {
                "id": str(id_val).strip(),
                "category": str(ws.cell(row=row, column=col_map["category"]).value or "").strip(),
                "label": str(ws.cell(row=row, column=col_map["label"]).value or "").strip(),
                "potenza_kw": float(ws.cell(row=row, column=col_map["potenza_kw"]).value or 0),
                "accumulo_kwh": float(ws.cell(row=row, column=col_map.get("accumulo_kwh", 0)).value or 0) if col_map.get("accumulo_kwh") else 0.0,
                "prezzo_eur": float(ws.cell(row=row, column=col_map["prezzo_eur"]).value or 0),
            }

            items.append(item)

        except Exception as e:
            errors.append(f"Row {row}: {str(e)}")

    if not items:
        raise HTTPException(status_code=400, detail="No valid items found in Excel file")

    # Check for duplicate IDs
    seen_ids = set()
    duplicates = []
    for item in items:
        if item["id"] in seen_ids:
            duplicates.append(item["id"])
        seen_ids.add(item["id"])

    if duplicates:
        raise HTTPException(
            status_code=400,
            detail=f"Duplicate IDs found: {', '.join(duplicates)}"
        )

    # Update metadata from Info sheet if present
    version = existing_catalog.get("version", "imported")
    source_pdf = existing_catalog.get("source_pdf", "")

    if "Info" in wb.sheetnames:
        ws_info = wb["Info"]
        if ws_info["B1"].value:
            version = str(ws_info["B1"].value)
        if ws_info["B2"].value:
            source_pdf = str(ws_info["B2"].value)

    # Build new catalog
    new_catalog = {
        "version": version,
        "source_pdf": source_pdf,
        "items": items
    }

    # Backup existing catalog
    if CATALOG_PATH.exists():
        backup_path = CATALOG_PATH.with_suffix(".json.backup")
        with open(CATALOG_PATH, "r", encoding="utf-8") as f:
            backup_content = f.read()
        with open(backup_path, "w", encoding="utf-8") as f:
            f.write(backup_content)

    # Save new catalog
    with open(CATALOG_PATH, "w", encoding="utf-8") as f:
        json.dump(new_catalog, f, indent=2, ensure_ascii=False)

    result = {
        "status": "success",
        "items_imported": len(items),
        "version": version
    }

    if errors:
        result["warnings"] = errors

    return result
